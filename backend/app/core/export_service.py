"""
Export Service - 导出服务
功能: 生成 XLSX 和 CSV 格式的分析报告（不依赖 pandas）

中文注释:
- 聚合 KPI、趋势、地理分布数据
- 生成多 Sheet 的 Excel 报告
- 生成单表 CSV 报告
"""

import csv
import io
from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook

if TYPE_CHECKING:
    from app.services.analytics_service import AnalyticsService


class ExportService:
    """
    分析报告导出服务
    """

    def __init__(self, analytics_service: "AnalyticsService"):
        """
        初始化导出服务

        中文注释: 依赖 AnalyticsService 获取数据
        """
        self.analytics_service = analytics_service

    async def generate_xlsx(self) -> io.BytesIO:
        """
        生成 XLSX 格式报告

        中文注释:
        - 多 Sheet 结构: KPI Summary, Trends, Geo Distribution
        - 直接使用 openpyxl（避免 pandas/numpy 依赖，显著加快部署构建）
        """
        output = io.BytesIO()

        # 获取所有数据
        kpi = await self.analytics_service.get_kpi_summary()
        trends = await self.analytics_service.get_submission_trends()
        geo = await self.analytics_service.get_author_geography()
        pipeline = await self.analytics_service.get_status_pipeline()

        wb = Workbook()

        def _write_table(ws, headers, rows):
            ws.append(list(headers))
            for r in rows:
                ws.append(list(r))

        # Sheet 1: KPI Summary
        ws_kpi = wb.active
        ws_kpi.title = "KPI Summary"
        _write_table(
            ws_kpi,
            headers=("指标", "数值"),
            rows=(
                ("本月新投稿", kpi.new_submissions_month),
                ("待处理稿件", kpi.total_pending),
                ("平均首次决定时间（天）", round(kpi.avg_first_decision_days, 1)),
                ("年度接受率", f"{kpi.yearly_acceptance_rate * 100:.1f}%"),
                ("本月 APC 收入", f"${kpi.apc_revenue_month:,.2f}"),
                ("年度 APC 收入", f"${kpi.apc_revenue_year:,.2f}"),
            ),
        )

        # Sheet 2: Submission Trends
        ws_trends = wb.create_sheet("Trends")
        if trends:
            _write_table(
                ws_trends,
                headers=("月份", "投稿数", "接受数"),
                rows=(
                    (
                        t.month.isoformat() if hasattr(t.month, "isoformat") else str(t.month),
                        t.submission_count,
                        t.acceptance_count,
                    )
                    for t in trends
                ),
            )
        else:
            _write_table(ws_trends, headers=("message",), rows=(("no data",),))

        # Sheet 3: Status Pipeline
        ws_pipeline = wb.create_sheet("Pipeline")
        if pipeline:
            _write_table(
                ws_pipeline,
                headers=("状态", "数量"),
                rows=((p.stage, p.count) for p in pipeline),
            )
        else:
            _write_table(ws_pipeline, headers=("message",), rows=(("no data",),))

        # Sheet 4: Geo Distribution
        ws_geo = wb.create_sheet("Geo Distribution")
        if geo:
            _write_table(
                ws_geo,
                headers=("国家", "投稿数"),
                rows=((g.country, g.submission_count) for g in geo),
            )
        else:
            _write_table(ws_geo, headers=("message",), rows=(("no data",),))

        # Sheet 5: Report Info
        ws_info = wb.create_sheet("Report Info")
        _write_table(
            ws_info,
            headers=("报告生成时间", "数据范围"),
            rows=((datetime.now().isoformat(), "过去 12 个月"),),
        )

        wb.save(output)
        output.seek(0)
        return output

    async def generate_csv(self) -> io.BytesIO:
        """
        生成 CSV 格式报告

        中文注释:
        - 单表结构，包含 KPI 汇总
        - UTF-8 编码
        """
        output = io.BytesIO()

        # 获取 KPI 数据
        kpi = await self.analytics_service.get_kpi_summary()

        rows = [
            ("本月新投稿", kpi.new_submissions_month),
            ("待处理稿件", kpi.total_pending),
            ("平均首次决定时间（天）", round(kpi.avg_first_decision_days, 1)),
            ("年度接受率", f"{kpi.yearly_acceptance_rate * 100:.1f}%"),
            ("本月 APC 收入", f"${kpi.apc_revenue_month:,.2f}"),
            ("年度 APC 收入", f"${kpi.apc_revenue_year:,.2f}"),
        ]

        text = io.StringIO()
        writer = csv.writer(text)
        writer.writerow(["指标", "数值"])
        writer.writerows(rows)
        output.write(text.getvalue().encode("utf-8"))
        output.seek(0)

        return output
